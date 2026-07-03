import pandas as pd
import os
import reportlab
import win32com.client as win32

from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter

import calendar
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet


VERSION = "1.7"

# =========================
# FOLDER SETUP
# =========================

REPORTS_FOLDER = "reports"
DATA_FOLDER = "data"
REPORT_EXPORTS_FOLDER = "reports exports"
KIT_RELEASE_FORMS_FOLDER = "release_forms"

# CSV reports from PowerShell app
SCANNED_FILE = os.path.join(REPORTS_FOLDER, "scanned_assets.csv")
MANUAL_FILE = os.path.join(REPORTS_FOLDER, "manual_assets.csv")

# Main Excel inventory file
INVENTORY_FILE = os.path.join(DATA_FOLDER, "inventory.xlsx")
REPORT_WORKBOOK = os.path.join(REPORTS_FOLDER, "asset_reports.xlsx")

# Create folders automatically
os.makedirs(REPORTS_FOLDER, exist_ok=True)
os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(REPORT_EXPORTS_FOLDER, exist_ok=True)
os.makedirs(KIT_RELEASE_FORMS_FOLDER, exist_ok=True)

# =========================
# VALID STATUSES
# =========================

VALID_STATUSES = {
    "Active",
    "Issued Out",
    "Returned",
    "Broken",
    "Retired"
}

def clean_filename(text):
    """Clean text to be used as a filename."""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        text = text.replace(char, "")
    return text.strip()

def generate_release_form(asset_row, issued_by="IT Department"):

    asset_id = asset_row["Asset ID"]
    asset_name = asset_row["AssetName"]
    user = asset_row["User"]
    issue_date = asset_row["Date"]
    asset_type = asset_row["AssetType"]
    model = asset_row["Model"]
    serial = asset_row["SerialNumber"]
    comments = asset_row["Comments"]
    

    filename = f"{KIT_RELEASE_FORMS_FOLDER}/"f"Kit Release Form-{clean_filename(user)}-{clean_filename(asset_name)}-{clean_filename(model)}-{asset_id}-{clean_filename(issue_date)}.pdf"

    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("IT Equipment Release Form", styles["Title"]))
    story.append(Spacer(1, 20))

    details = [
        ["Asset ID", str(asset_id)],
        ["Item", str(asset_type)],
        ["Description", str(asset_name)],
        ["Model Number", str(model)],
        ["Serial Number", str(serial)],
        ["Issued To", str(user)],
        ["Issued By", str(issued_by)],
        ["Issued On", str(issue_date)],
        ["Comments", str(comments)],
    ]

    table = Table(details, colWidths=[140, 300])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    story.append(table)
    story.append(Spacer(1, 40))

    agreement = """
    I confirm that I have received the equipment listed above in good condition
    and accept responsibility for its care and return. I understand that I am responsible for any loss or damage to the equipment while it is in my possession.
    """

    story.append(Paragraph(agreement, styles["BodyText"]))
    story.append(Spacer(1, 50))

    story.append(Paragraph("Employee Signature: ____________________", styles["BodyText"]))
    story.append(Spacer(1, 20))
    story.append(Paragraph("Date: ____________________", styles["BodyText"]))

    doc.build(story)

    return filename


def send_email_asset_issue(user_email, user_name, asset_id, asset_name, asset_model, shared_email, asset_type):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = user_email
    mail.CC = shared_email
    mail.Subject = f"Asset Issued: {asset_type}, {asset_name} {asset_model}"
    mail.Body = f"Hi {user_name},\n\nYou have been issued the following asset:\n\nItem: {asset_name} {asset_model}\nAsset Type: {asset_type}\n\nPlease contact IT if you have any issues.\n\nKind regards,\nIT Department"
    mail.Send()

def send_email_asset_return(user_email, user_name, asset_id, asset_name, asset_model, shared_email, asset_type):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = user_email
    mail.CC = shared_email
    mail.Subject = f"Asset Returned: {asset_type}, {asset_name} {asset_model}"
    mail.Body = f"Hi {user_name},\n\nYou have returned the following asset:\n\nName: {asset_name} {asset_model}\nAsset Type: {asset_type}\n\nThank you for returning the asset.\n\nKind regards,\nIT Department"
    mail.Send()
# =========================
# SAVE / LOAD FUNCTIONS
# =========================

def save_inventory(df):
    """
    Save inventory while preserving:
    - filters
    - formatting
    - tables
    - column widths
    - colors
    """

    # Create workbook if missing
    if not os.path.exists(INVENTORY_FILE):

        with pd.ExcelWriter(
            INVENTORY_FILE,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="Assets",
                index=False
            )

        return

    # Load existing workbook
    workbook = load_workbook(INVENTORY_FILE)

    # Ensure Assets sheet exists
    if "Assets" not in workbook.sheetnames:

        sheet = workbook.create_sheet("Assets")

        # Add headers
        sheet.append(list(df.columns))

    else:
        sheet = workbook["Assets"]

    # Delete old data rows only
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

    # Write updated rows
    for row in df.itertuples(index=False):

        sheet.append(list(row))

    # Save workbook
    workbook.save(INVENTORY_FILE)


def repair_asset(asset_id):

    df = load_inventory()

    if df is None:
        return

    asset_id = str(asset_id).strip()
    match = (df["Asset ID"].astype(str).str.strip() == asset_id)

    if not match.any():

        print("Asset not found")
        pause()
        return

    current_status = df.loc[match, "Status"].iloc[0]

    if current_status != "Broken":

        print(f"Asset {asset_id} is not marked as Broken, only Broken assets can be repaired.")
        pause()
        return

    repair_notes = input("Enter repair notes: ")

    old_status = current_status

    current_user = ""
    

    if "User" in df.columns:
        current_user = df.loc[match, "User"].iloc[0]

    df.loc[match, "Status"] = "Active"

    save_inventory(df)

    log_history(
        asset_id,
        old_status,
        "Active",
        current_user
    )

    add_asset_log(
        asset_id,
        "Repaired",
        repair_notes
    )

    print(f"{asset_id} repaired and set to Active")

    pause()



def add_asset_log(asset_id, log_type, notes):
    if not os.path.exists(INVENTORY_FILE):
        return
    
    workbook = load_workbook(INVENTORY_FILE)

    if "Asset Logs" not in workbook.sheetnames:
        log_sheet = workbook.create_sheet("Asset Logs")

        log_sheet.append(["Timestamp", "Asset ID", "Asset Name", "Serial Number", "Log Type", "Notes"])
    else:
        log_sheet = workbook["Asset Logs"]

    df = load_inventory()
    asset_name = ""
    serial_number = ""

    if df is not None:
        asset_row = df[df["Asset ID"] == asset_id]

        if not asset_row.empty:
            if "AssetName" in asset_row.columns:
                asset_name = asset_row["AssetName"].iloc[0]
            if "SerialNumber" in asset_row.columns:
                serial_number = asset_row["SerialNumber"].iloc[0]

    log_sheet.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        asset_id,
        asset_name,
        serial_number,
        log_type,
        notes
    ])
    workbook.save(INVENTORY_FILE)

    print(f"Log added for {asset_id}: {log_type} - {notes}")

def view_asset_logs(asset_id):
    if not os.path.exists(INVENTORY_FILE):
        print("No logs found.")
        return
    workbook = load_workbook(INVENTORY_FILE)
    if "Asset Logs" not in workbook.sheetnames:
        print("No logs found.")
        return
    
    log_sheet = workbook["Asset Logs"]
    print(f"\nLogs for {asset_id}:\n")
    found = False

    for row in log_sheet.iter_rows(min_row=2, values_only=True):
        timestamp, log_asset_id, log_asset_name, log_serial_number, log_type, notes = row

        if str(log_asset_id) == str(asset_id):
            print(f"{timestamp} - {log_type} - {notes}")
            found = True

    if not found:
        print("No logs found.")

    pause()

def monthly_summary_report():
    try:
        month = int(input("Enter month (1-12): ").strip())
        year = int(input("Enter year (e.g., 2024): ").strip())
    except ValueError:
        print("Invalid month or year")
        pause()
        return
    
    history = pd.read_excel(INVENTORY_FILE, sheet_name="History")

    history['Timestamp'] = pd.to_datetime(history['Timestamp'], errors='coerce')
    
    report = history[(history['Timestamp'].dt.month == month) & (history['Timestamp'].dt.year == year)]
    
    if report.empty:
        print(f"No history found for {month}/{year}")
        pause()
        return
    
    issued = (report['New Status'] == 'Issued Out').sum()
    returned = (report['New Status'] == 'Returned').sum()
    broken = (report['New Status'] == 'Broken').sum()
    retired = (report['New Status'] == 'Retired').sum()

    print(f"\nMonthly Summary for {month}/{year}:")
    print(f"Issued Out: {issued}")
    print(f"Returned: {returned}")
    print(f"Broken: {broken}")
    print(f"Retired: {retired}")
    
    print("\nDetailed Changes:")
    print(report[['Timestamp', 'Asset ID', 'Old Status', 'New Status', 'User', 'Email']].to_string(index=False))

    save_pdf = input("Export as PDF? (y/n): ").strip().lower()

    if save_pdf == "y":
        export_monthly_report_pdf(report, month, year)

    pause()

def log_history(asset_id, old_status, new_status, user = "", email = ""):
    if not os.path.exists(INVENTORY_FILE):
        return
    
    workbook = load_workbook(INVENTORY_FILE)

    if "History" not in workbook.sheetnames:
        history_sheet = workbook.create_sheet("History")

        history_sheet.append(["Timestamp", "Asset ID", "Old Status", "New Status", "User", "Email"])

    else:
        history_sheet = workbook["History"]

    history_sheet.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        asset_id,
        old_status,
        new_status,
        user,
        email
    ])
    workbook.save(INVENTORY_FILE)

def export_monthly_report_pdf(report, month, year):
    month_name = calendar.month_name[month]
    pdf_path = os.path.join(REPORT_EXPORTS_FOLDER, f"{month_name}_{year}_report.pdf")

    doc = SimpleDocTemplate(pdf_path)
    styles = getSampleStyleSheet()
    story = []

    issued = (report['New Status'] == 'Issued Out').sum()
    returned = (report['New Status'] == 'Returned').sum()
    broken = (report['New Status'] == 'Broken').sum()
    retired = (report['New Status'] == 'Retired').sum()

    story.append(Paragraph(f"IT Asset Report - {month_name} {year}", styles["Title"]))
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 20))

    story.append(Paragraph(f"Total Actions: {len(report)}", styles["Normal"]))
    story.append(Paragraph(f"Issued Out: {issued}", styles["Normal"]))
    story.append(Paragraph(f"Returned: {returned}", styles["Normal"]))
    story.append(Paragraph(f"Broken: {broken}", styles["Normal"]))
    story.append(Paragraph(f"Retired: {retired}", styles["Normal"]))

    story.append(Spacer(1, 20))
    story.append(Paragraph("Detailed Activity", styles["Heading2"]))
    story.append(Spacer(1, 10))

    for _, row in report.iterrows():
        line = (
            f"Timestamp: {row['Timestamp']}|"
            f"Asset ID: {row['Asset ID']}|"
            f"Status: {row['Old Status']} -> {row['New Status']}|"
            f"User: {row['User']}|"
        )
        story.append(Paragraph(line, styles["Normal"]))
        story.append(Spacer(1, 8))

    doc.build(story)

    print(f"PDF saved to: {pdf_path}")




def load_inventory():
    """Load inventory dataframe from Excel."""

    if not os.path.exists(INVENTORY_FILE):
        print("Inventory file not found.")
        return None

    return pd.read_excel(INVENTORY_FILE)


# =========================
# CREATE / UPDATE INVENTORY
# =========================



def find_asset_match(inventory_df, asset):

    asset_id = str(asset.get("Asset ID", "")).strip()
    serial = str(asset.get("SerialNumber", "")).strip()
    mac = str(asset.get("MACAddress", "")).strip()

    # Match by Asset ID
    if asset_id:
        matches = inventory_df[
            inventory_df["Asset ID"].astype(str).str.strip() == asset_id
        ]
        if not matches.empty:
            return matches.index[0]

    # Match by Serial Number
    if serial and serial != "N/A":
        matches = inventory_df[
            inventory_df["SerialNumber"].astype(str).str.strip() == serial
        ]
        if not matches.empty:
            return matches.index[0]

    # Match by MAC Address
    if mac:
        matches = inventory_df[
            inventory_df["MACAddress"].astype(str).str.strip() == mac
        ]
        if not matches.empty:
            return matches.index[0]

    return None


def update_inventory():

    # Load CSV source files
    scanned_df = pd.read_csv(SCANNED_FILE)
    manual_df = pd.read_csv(MANUAL_FILE)

    # Combine import files
    import_df = pd.concat(
        [scanned_df, manual_df],
        ignore_index=True
    )


    # Load existing inventory
    inventory_df = load_inventory()

    if inventory_df is None:

        inventory_df = pd.DataFrame()


    # Process imported assets
    for _, asset in import_df.iterrows():

        match_index = None

        if not inventory_df.empty:

            match_index = find_asset_match(
                inventory_df,
                asset
            )

        # Existing asset found
        if match_index is not None:

            fields_to_update = [

                "AssetName",
                "AssetType",
                "Model",
                "SerialNumber",
                "MACAddress",
                "Comments",
                "IMEI"

            ]

            for field in fields_to_update:

                if field in inventory_df.columns:

                    inventory_df.at[
                        match_index,
                        field
                    ] = asset.get(field)

        # New asset
        else:

            new_asset = asset.to_dict()

            if "Status" not in new_asset:
                new_asset["Status"] = "Active"

            if "Asset ID" not in new_asset:
                new_asset["Asset ID"] = ""

            inventory_df = pd.concat(
                [
                    inventory_df,
                    pd.DataFrame([new_asset])
                ],
                ignore_index=True
            )

    # Ensure Status column exists
    if "Status" not in inventory_df.columns:
        inventory_df["Status"] = "Active"

    # Ensure Asset ID column exists
    if "Asset ID" not in inventory_df.columns:
        inventory_df["Asset ID"] = ""

    save_inventory(inventory_df)

    print("Inventory updated successfully.")

    pause()
    clear_screen()



# =========================
# UPDATE ASSET STATUS
# =========================

def update_asset_status(asset_id, new_status):

    df = load_inventory()

    if df is None:
        return

    current_user = ""

    if new_status not in VALID_STATUSES:

        print(
            f"Invalid status. Use: "
            f"{', '.join(VALID_STATUSES)}"
        )

        return

    asset_id = str(asset_id).strip()

    match = (
        df["Asset ID"].astype(str).str.strip() == asset_id
    )

    if not match.any():

        print(f"Asset ID {asset_id} not found.")
        pause()
        return

    old_status = df.loc[match, "Status"].iloc[0]

    if "User" in df.columns:
        current_user = df.loc[match, "User"].iloc[0]

    df.loc[match, "Status"] = new_status

    save_inventory(df)

    log_history(
        asset_id,
        old_status,
        new_status,
        current_user
    )

    print(f"{asset_id} updated to {new_status}")

    pause()
    clear_screen()



# =========================
# SEARCH FUNCTIONS
# =========================

def search_assets_by_serial(serial):

    df = load_inventory()

    if df is None:
        return

    result = df[df["SerialNumber"].astype(str).str.contains(
        serial,
        case=False,
        na=False
    )]

    if result.empty:
        print("No matching serial numbers found")
    else:
        for _, row in result.iterrows():
            print("=" * 50)
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)

    pause()
    clear_screen()


def search_assets_by_asset_name(asset_name):

    df = load_inventory()

    if df is None:
        return

    result = df[df["AssetName"].astype(str).str.contains(
        asset_name,
        case=False,
        na=False
    )]

    if result.empty:
        print("No matching assets found")
    else:
        for _, row in result.iterrows():
            print("=" * 50)
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)

    pause()
    clear_screen()


def search_assets_by_asset_type(asset_type):

    df = load_inventory()

    if df is None:
        return

    result = df[df["AssetType"].astype(str).str.contains(
        asset_type,
        case=False,
        na=False
    )]

    if result.empty:
        print("No matching asset types found")
    else:
        for _, row in result.iterrows():
            print("=" * 50)
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)
    pause()
    clear_screen()


def search_assets_by_model(model):

    df = load_inventory()

    if df is None:
        return

    result = df[df["Model"].astype(str).str.contains(
        model,
        case=False,
        na=False
    )]

    if result.empty:
        print("No matching models found")
    else:
        print("=" * 50)
        for _, row in result.iterrows():
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)

    pause()
    clear_screen()


def search_asset(asset_id):

    df = load_inventory()

    if df is None:
        return

    result = df[df["Asset ID"].astype(str).str.strip() == str(asset_id).strip()]

    if result.empty:
        print(f"No asset found with ID {asset_id}")
    else:
        print("=" * 50)
        for _, row in result.iterrows():
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)

    pause()
    clear_screen()


def search_assets_by_person(person):

    df = load_inventory()

    if df is None:
        return

    result = df[df["User"].astype(str).str.contains(
        person,
        case=False,
        na=False
    )]

    if result.empty:
        print(f"No assets found for {person}")
    else:
        print("=" * 50)
        for _, row in result.iterrows():
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)

    pause()
    clear_screen()


def search_assets_by_location(location):

    df = load_inventory()

    if df is None:
        return

    result = df[df["Location"].astype(str).str.contains(
        location,
        case=False,
        na=False
    )]

    if result.empty:
        print(f"No assets found at {location}")
    else:
        print("=" * 50)
        for _, row in result.iterrows():
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)

    pause()
    clear_screen()


def search_assets_by_status(status):

    df = load_inventory()

    if df is None:
        return

    result = df[df["Status"].astype(str).str.contains(
        status,
        case=False,
        na=False
    )]

    if result.empty:
        print(f"No assets found with status {status}")
    else:
        print("=" * 50)
        for _, row in result.iterrows():
            print(f"Asset ID: {row['Asset ID']}") 
            print(f"Asset Name: {row['AssetName']}")
            print(f"Serial Number: {row['SerialNumber']}")
            print(f"Asset Type: {row['AssetType']}")
            print(f"Model: {row['Model']}")
            print(f"Status: {row['Status']}")
            print(f"User: {row['User']}")
            print(f"Location: {row['Location']}")
            print("=" * 50)

    pause()
    clear_screen()


# =========================
# ISSUE OUT ASSET
# =========================

def issue_out_asset(asset_id, person, location):

    df = load_inventory()

    if df is None:
        return

    asset_id = str(asset_id).strip()
    match = df["Asset ID"].astype(str).str.strip() == asset_id

    if not match.any():
        print("Asset not found")
        pause()
        return

    if match.sum() > 1:
        print("Error: Duplicate Asset IDs found")
        pause()
        return

    current_status = str(df.loc[match, "Status"].iloc[0]).strip()
    current_user = str(df.loc[match, "User"].iloc[0]).strip()
    current_location = str(df.loc[match, "Location"].iloc[0]).strip()

    if (current_user and current_user.lower() != "nan" and current_location != "IT Storage"):
        print(f"Asset already issued to {current_user}")
        pause()
        return

    if current_status in {"Broken", "Retired"}:
        print(f"Cannot issue asset because status is {current_status}")
        pause()
        return

    old_status = current_status

    df.loc[match, "Status"] = "Active"
    df.loc[match, "User"] = person
    df.loc[match, "Location"] = location

    save_inventory(df)

    asset = df.loc[match].iloc[0]
    pdf_path = generate_release_form(asset, issued_by="IT Department")
    print (f"Release form generated: {pdf_path}")

    email = input("Enter user email for notification: ").strip()

    send_email_asset_issue(
        user_email=email,
        user_name=person,
        asset_id=asset_id,
        asset_name=df.loc[match, "AssetName"].iloc[0],
        asset_model=df.loc[match, "Model"].iloc[0],
        asset_type=df.loc[match, "AssetType"].iloc[0],
        shared_email="hulme1905@outlook.com"
    )

    log_history(
        asset_id,
        old_status,
        "Issued Out",
        person,
        email
    )

    print(f"{asset_id} issued to {person}")

    pause()
    clear_screen()



# =========================
# RETURN ASSET
# =========================

def return_asset(asset_id):

    df = load_inventory()

    if df is None:
        return

    asset_id = str(asset_id).strip()
    match = df["Asset ID"].astype(str).str.strip() == asset_id

    if not match.any():
        print("Asset not found")
        pause()
        return
    
    old_status = df.loc[match, "Status"].iloc[0]

    current_user = df.loc[match, "User"].iloc[0]

    df.loc[match, "Status"] = "Returned"

    df.loc[match, "User"] = ""

    df.loc[match, "Location"] = "IT Storage"

    save_inventory(df)

    email = input("Enter user email for notification: ").strip()

    send_email_asset_return(
        user_email=email,
        user_name=current_user,
        asset_id=asset_id,
        asset_name=df.loc[match, "AssetName"].iloc[0],
        asset_model=df.loc[match, "Model"].iloc[0],
        asset_type=df.loc[match, "AssetType"].iloc[0],
        shared_email="hulme1905@outlook.com")

    log_history(
        asset_id, 
        old_status, 
        "Returned", 
        current_user,
        email
    )

    print(f"{asset_id} returned to IT Storage")

    pause()
    clear_screen()


# =========================
# MARK AS BROKEN
# =========================

def mark_asset_broken(asset_id):

    df = load_inventory()

    if df is None:
        return

    asset_id = str(asset_id).strip()
    match = df["Asset ID"].astype(str).str.strip() == asset_id

    if not match.any():
        print("Asset not found")
        pause()
        return

    old_status = df.loc[match, "Status"].iloc[0]

    current_user = df.loc[match, "User"].iloc[0]

    df.loc[match, "Status"] = "Broken"

    fault_description = input("Enter fault description: ")

    save_inventory(df)

    log_history(
        asset_id,
        old_status,
        "Broken", 
        current_user
    )

    add_asset_log(
        asset_id,
        "Fault Reported",
        fault_description
    )

    print(f"{asset_id} marked as Broken")

    pause()
    clear_screen()


# =========================
# RETIRE ASSET
# =========================

def retire_asset(asset_id):

    df = load_inventory()

    if df is None:
        return

    asset_id = str(asset_id).strip()
    match = df["Asset ID"].astype(str).str.strip() == asset_id

    if not match.any():
        print("Asset not found")
        pause()
        return
    
    old_status = df.loc[match, "Status"].iloc[0]

    current_user = df.loc[match, "User"].iloc[0]

    df.loc[match, "Status"] = "Retired"

    save_inventory(df)

    log_history(
        asset_id,
        old_status,
        "Retired", 
        current_user
    )

    print(f"{asset_id} retired")

    pause()
    clear_screen()


def generate_reports():

    df = load_inventory()

    if df is None:
        return

    if os.path.exists(REPORT_WORKBOOK):

        workbook = load_workbook(REPORT_WORKBOOK)

    else:

        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    report_sheets = [

        "Summary",
        "Assigned Assets",
        "Broken Assets",
        "Returned Assets",
        "Retired Assets",
        "IT Storage Assets",
        "User Assets"

    ]

    # Delete old report sheets

    for sheet_name in report_sheets:

        if sheet_name in workbook.sheetnames:

            del workbook[sheet_name]

    # =====================
    # SUMMARY
    # =====================

    summary_sheet = workbook.create_sheet("Summary")

    summary_sheet.append(["Metric", "Count"])

    summary_sheet.append(["Total Assets", len(df)])

    summary_sheet.append([
        "Active",
        len(df[df["Status"] == "Active"])
    ])

    summary_sheet.append([
        "Broken",
        len(df[df["Status"] == "Broken"])
    ])

    summary_sheet.append([
        "Returned",
        len(df[df["Status"] == "Returned"])
    ])

    summary_sheet.append([
        "Retired",
        len(df[df["Status"] == "Retired"])
    ])

    assigned_count = len(

        df[
            (df["Status"] == "Active")
            &
            (df["User"].notna())
            &
            (df["User"] != "")
        ]

    )

    summary_sheet.append([
        "Assigned Assets",
        assigned_count
    ])

    # =====================
    # ASSIGNED ASSETS
    # =====================

    assigned_df = df[
        (df["Status"] == "Active")
        &
        (df["User"].notna())
        &
        (df["User"] != "")
    ]

    assigned_sheet = workbook.create_sheet(
        "Assigned Assets"
    )

    assigned_sheet.append(
        list(assigned_df.columns)
    )

    for row in assigned_df.itertuples(index=False):

        assigned_sheet.append(list(row))

    # =====================
    # BROKEN ASSETS
    # =====================

    broken_df = df[
        df["Status"] == "Broken"
    ]

    broken_sheet = workbook.create_sheet(
        "Broken Assets"
    )

    broken_sheet.append(
        list(broken_df.columns)
    )

    for row in broken_df.itertuples(index=False):

        broken_sheet.append(list(row))

    # =====================
    # RETURNED ASSETS
    # =====================

    returned_df = df[
        df["Status"] == "Returned"
    ]

    returned_sheet = workbook.create_sheet(
        "Returned Assets"
    )

    returned_sheet.append(
        list(returned_df.columns)
    )

    for row in returned_df.itertuples(index=False):

        returned_sheet.append(list(row))

    # =====================
    # RETIRED ASSETS
    # =====================

    retired_df = df[
        df["Status"] == "Retired"
    ]

    retired_sheet = workbook.create_sheet(
        "Retired Assets"
    )

    retired_sheet.append(
        list(retired_df.columns)
    )

    for row in retired_df.itertuples(index=False):

        retired_sheet.append(list(row))

    # =====================
    # IT STORAGE
    # =====================

    storage_df = df[
        df["Location"] == "IT Storage"
    ]

    storage_sheet = workbook.create_sheet(
        "IT Storage Assets"
    )

    storage_sheet.append(
        list(storage_df.columns)
    )

    for row in storage_df.itertuples(index=False):

        storage_sheet.append(list(row))

    # =====================
    # USER ASSETS
    # =====================

    user_assets_df = df[
        (df["User"].notna())
        &
        (df["User"] != "")
    ]

    user_assets_df = user_assets_df.sort_values(
        by=["User", "AssetName"]
    )

    user_sheet = workbook.create_sheet(
        "User Assets"
    )

    user_sheet.append(
        list(user_assets_df.columns)
    )

    for row in user_assets_df.itertuples(index=False):

        user_sheet.append(list(row))

    # Add filters and freeze top row

    for sheet in workbook.worksheets:

        if sheet.max_row > 1:

            sheet.auto_filter.ref = sheet.dimensions
            sheet.freeze_panes = "A2"
            sheet.protection.sheet = True
            sheet.protection.autoFilter = False
            sheet.protection.sort = False


            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save(REPORT_WORKBOOK)

    print("Reports generated successfully.")

    pause()
    clear_screen()


def menu():

    while True:

        clear_screen()
        show_dashboard()

        print("[1.] Update Inventory")
        print("[2.] Search Asset")
        print("[3.] Issue Asset")
        print("[4.] Return Asset")
        print("[5.] Mark Asset Broken")
        print("[6.] Retire Asset")
        print("[7.] View Asset Logs")
        print("[8.] Repair Asset")
        print("[9.] Reports Menu")
        print("[10.] Exit")
        print("=" * 50)

        choice = input("\nSelect option: ")

        if choice == "1":

            update_inventory()

        elif choice == "2":

            search_menu()


        elif choice == "3":

            asset_id = input("Asset ID: ")
            person = input("User: ")
            location = input("Location: ")

            issue_out_asset(
                asset_id,
                person,
                location
            )

        elif choice == "4":

            asset_id = input("Asset ID: ")
            return_asset(asset_id)

        elif choice == "5":

            asset_id = input("Asset ID: ")
            mark_asset_broken(asset_id)

        elif choice == "6":

            asset_id = input("Asset ID: ")
            retire_asset(asset_id)

        elif choice == "7":

            asset_logs_menu()

        elif choice == "8":

            asset_id = input("Asset ID: ")
            repair_asset(asset_id)

        elif choice == "9":

            report_menu()

        elif choice == "10":

            print("Goodbye")
            break
        else:

            print("Invalid option")


def search_menu():

    while True:

        clear_screen()
        show_dashboard()
        
        print("\n===== Asset Search =====")
        print("=" * 50)

        print("[1.] Asset ID")
        print("[2.] User")
        print("[3.] Location")
        print("[4.] Status")
        print("[5.] Serial Number")
        print("[6.] Asset Name")
        print("[7.] Model")
        print("[8.] Asset Type")
        print("[9.] Back")
        print("=" * 50)

        choice = input("Select option: ")

        if choice == "1":

            search_asset(
                input("Asset ID: ")
            )

        elif choice == "2":

            search_assets_by_person(
                input("User: ")
            )

        elif choice == "3":

            search_assets_by_location(
                input("Location: ")
            )

        elif choice == "4":

            search_assets_by_status(
                input("Status: ")
            )

        elif choice == "5":

            search_assets_by_serial(
                input("Serial Number: ")
            )

        elif choice == "6":

            search_assets_by_asset_name(
                input("Asset Name: ")
            )

        elif choice == "7":

            search_assets_by_model(
                input("Model: ")
            )

        elif choice == "8":

            search_assets_by_asset_type(
                input("Asset Type: ")
            )

        elif choice == "9":

            break

        else:

            print("Invalid option")

def asset_logs_menu():
    while True:
        clear_screen()
        show_dashboard()
        print("=" * 50)
        print("\n===== View Asset Logs =====")
        print("[1.] Add log entry")
        print("[2.] View logs for asset")
        print("[3.] Back")
        print("=" * 50)

        choice = input("\nSelect option: ")

        if choice == "1":
            asset_id = input("Asset ID: ")
            log_type = input("Log Type (e.g. Maintenance, Repair, Note): ")
            notes = input("Notes: ")
            add_asset_log(asset_id, log_type, notes)
            pause()
        elif choice == "2":
            asset_id = input("Asset ID: ")
            view_asset_logs(asset_id)
        elif choice == "3":
            break
        else:
            print("Invalid option")
            pause()

def report_menu():
    while True:
        clear_screen()
        print("=" * 50)
        print("\n===== Reports Menu =====")
        print("[1.] Generate Reports")
        print("[2.] Monthly Summary Report")
        print("[3.] Back")
        print("=" * 50)

        choice = input("\nSelect option: ")

        if choice == "1":
            generate_reports()
        elif choice == "2":
            monthly_summary_report()
        elif choice == "3":
            break
        else:
            print("Invalid option")
            pause()

def pause():
    input("\nPress Enter to continue...")

def clear_screen():

    os.system(
        "cls" if os.name == "nt" else "clear"
    )

def show_dashboard():

    df = load_inventory()

    if df is None:
        print("No inventory data found.")
        return
    
    total_assets = len(df)
    active_assets = len(df[df["Status"] == "Active"])
    assets_in_storage = len(df[df["Location"] == "IT Storage"])
    returned_assets = len(df[df["Status"] == "Returned"])
    broken_assets = len(df[df["Status"] == "Broken"])
    retired_assets = len(df[df["Status"] == "Retired"])

    print("=" * 50)
    print(f"\n===== IT Inventory Management System {VERSION} =====")
    print("=" * 50)
    print(f"Total Assets: {total_assets}")
    print(f"Active Assets: {active_assets}")
    print(f"Assets in Storage: {assets_in_storage}")
    print(f"Returned Assets: {returned_assets}")
    print(f"Broken Assets: {broken_assets}")
    print(f"Retired Assets: {retired_assets}")
    print("=" * 50)



# =========================
# MAIN
# =========================

if __name__ == "__main__":
    menu()